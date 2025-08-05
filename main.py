import base64
import json
import os
import re
import sqlite3
import time

import pandas as pd
from curl_cffi import requests
from openpyxl import load_workbook
from tqdm import tqdm

from settings import *


class SQLiteDeduplicator:
    def __init__(self, db_filepath="deduplicator.db"):
        """
        Initializes the SQLite deduplicator for combined URL and JSON data.

        Args:
            db_filepath (str): The path to the SQLite database file.
        """
        self.db_filepath = db_filepath
        self._create_table()

    def _create_table(self):
        """
        Creates the 'records' table in the database if it doesn't exist.
        The 'dedup_key' column has a UNIQUE constraint to ensure deduplication.
        """
        conn = None
        try:
            conn = sqlite3.connect(self.db_filepath)
            cursor = conn.cursor()
            cursor.execute(
                '''
                CREATE TABLE IF NOT EXISTS records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    dedup_key TEXT UNIQUE NOT NULL
                )
            '''
            )
            conn.commit()
            # 为 dedup_key 列创建索引以提高查询速度
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_dedup_key ON records (dedup_key)')
            conn.commit()
        except sqlite3.Error as e:
            print(f"Database initialization error: {e}")
        finally:
            if conn:
                conn.close()

    def _generate_dedup_key(self, url, json_data):
        """
        Generates a unique deduplication key by combining URL and normalized JSON data.

        Args:
            url (str): The URL string.
            json_data (dict or None): The dictionary for JSON data, or None if no JSON data.

        Returns:
            str: The combined deduplication key.
        """
        # 1. 规范化 JSON 数据：确保相同内容的 JSON 无论键顺序如何，都生成相同的字符串
        if json_data is not None:
            # 使用 separators 和 sort_keys 保证一致性
            normalized_json_str = json.dumps(json_data, separators=(',', ':'), sort_keys=True)
        else:
            normalized_json_str = ""  # 如果没有 JSON 数据，则为空字符串

        # 2. 组合 URL 和规范化后的 JSON 字符串
        # 我们使用一个不太可能出现在URL或JSON数据中的分隔符，
        # 确保组合后的字符串是唯一的。例如，'\u001F' (Unit Separator)
        combined_key = f"{url}\u001F{normalized_json_str}"
        return combined_key

    def check_duplicate(self, url, json_data=None):
        """
        Checks if a record (based on URL + JSON data) exists in the database.

        Args:
            url (str): The URL of the record.
            json_data (dict, optional): The associated JSON data. Defaults to None.

        Returns:
            bool: True if the record is found, False otherwise.
        """
        dedup_key = self._generate_dedup_key(url, json_data)
        conn = None
        try:
            conn = sqlite3.connect(self.db_filepath)
            cursor = conn.cursor()
            cursor.execute("SELECT 1 FROM records WHERE dedup_key = ?", (dedup_key,))
            return cursor.fetchone() is not None
        except sqlite3.Error as e:
            print(f"Database query error for key '{dedup_key}': {e}")
            return False
        finally:
            if conn:
                conn.close()

    def add_record(self, url, json_data=None):
        """
        Adds a new record (based on URL + JSON data) to the database.

        Args:
            url (str): The URL of the record.
            json_data (dict, optional): The associated JSON data. Defaults to None.

        Returns:
            bool: True if the record was successfully added (not a duplicate), False otherwise.
        """
        dedup_key = self._generate_dedup_key(url, json_data)
        conn = None
        try:
            conn = sqlite3.connect(self.db_filepath)
            cursor = conn.cursor()
            cursor.execute("INSERT INTO records (dedup_key) VALUES (?)", (dedup_key,))
            conn.commit()
            return True  # Record successfully added
        except sqlite3.IntegrityError:
            return False  # Duplicate record
        except sqlite3.Error as e:
            print(f"Database insert error for key '{dedup_key}': {e}")
            return False
        finally:
            if conn:
                conn.close()

    def get_record_count(self):
        """
        Returns the total number of unique records in the database.
        """
        conn = None
        try:
            conn = sqlite3.connect(self.db_filepath)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM records")
            return cursor.fetchone()[0]
        except sqlite3.Error as e:
            print(f"Database count error: {e}")
            return 0
        finally:
            if conn:
                conn.close()


def read_excel_to_dict_list(file_path):
    df = pd.read_excel(file_path)
    records = df.to_dict(orient='records')
    return records


def write_list_to_excel(new_data, headers, target_file):
    if not os.path.exists(target_file):
        # 文件不存在，创建新文件并写入表头
        df = pd.DataFrame(new_data, columns=headers)
        df.to_excel(target_file, index=False)
        print(f"已新建文件 {target_file} 并写入数据 >>> {new_data[0]}")
    else:
        # 文件已存在，追加数据（不写表头）
        df_new = pd.DataFrame(new_data, columns=headers)
        with pd.ExcelWriter(
            target_file, mode='a', engine='openpyxl', if_sheet_exists='overlay'
        ) as writer:
            # 找到当前已有数据行数（避免覆盖）
            book = load_workbook(target_file)
            sheet = book.active
            start_row = sheet.max_row + 1
            df_new.to_excel(writer, index=False, header=False, startrow=start_row - 1)
        print(f"已向 {target_file} 追加写入数据 >>> {new_data[0]}")


def write_list_to_csv(new_data, headers, target_file):
    df = pd.DataFrame(new_data, columns=headers)

    # 判断文件是否存在
    file_exists = os.path.isfile(target_file)

    # 写入或追加
    df.to_csv(target_file, mode='a', index=False, header=not file_exists, encoding='utf-8-sig')

    if file_exists:
        print(f"已向 {target_file} 追加写入数据 >>> {new_data[0][1][:50]} ......")
    else:
        print(f"已新建文件 {target_file} 并写入数据 >>> {new_data[0][1][:50]} ......")


def get_commentable_id(url, session):
    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'dnt': '1',
        'pragma': 'no-cache',
        'priority': 'u=0, i',
        'upgrade-insecure-requests': '1',
        'user-agent': random.choice(CHROME_USER_AGENTS),
    }

    for i in range(5):
        try:
            response = session.get(url, headers=headers)
            commentable_id_match = re.search(r'data-commentable_id="(.*?)"', response.text)
            if commentable_id_match:
                commentable_id = commentable_id_match.group(1)
                print("成功提取到commentable_id：", commentable_id)
            else:
                raise Exception("未匹配到commentable_id")

            csrf_token_match = re.compile(r'<meta name="csrf-token" content="(.*?)"').search(
                response.text
            )
            if csrf_token_match:
                csrf_token = csrf_token_match.group(1)
                print(f"匹配到的 CSRF Token 为: {csrf_token}")
            else:
                raise Exception("未找到匹配的 CSRF Token。")

            return commentable_id, csrf_token
        except Exception as e:
            session.impersonate = random.choice(CURL_CFFI_BROWSER)
            print('提取commentable_id失败，正在重试中 >>> ', e)


def get_comments(url, commentable_id, session, csrf_token, next_cursor=None, page=1, file_name=''):
    json_data = [
        {
            'operationName': 'CommentsQuery',
            'variables': {
                'commentableId': commentable_id,
                'nextCursor': next_cursor,
                'previousCursor': None,
                'first': 25,
                'last': None,
            },
            'query': 'query CommentsQuery($commentableId: ID!, $nextCursor: String, $previousCursor: String, $replyCursor: String, $first: Int, $last: Int) {\n  commentable: node(id: $commentableId) {\n    id\n    ... on Project {\n      url\n      __typename\n    }\n    ... on Commentable {\n      canComment\n      canCommentSansRestrictions\n      commentsCount\n      projectRelayId\n      canUserRequestUpdate\n      comments(\n        first: $first\n        last: $last\n        after: $nextCursor\n        before: $previousCursor\n      ) {\n        edges {\n          node {\n            ...CommentInfo\n            ...CommentReplies\n            __typename\n          }\n          __typename\n        }\n        pageInfo {\n          startCursor\n          hasNextPage\n          hasPreviousPage\n          endCursor\n          __typename\n        }\n        __typename\n      }\n      __typename\n    }\n    __typename\n  }\n  me {\n    id\n    name\n    imageUrl(width: 200)\n    isKsrAdmin\n    url\n    isBlocked\n    userRestrictions {\n      restriction\n      releaseAt\n      __typename\n    }\n    __typename\n  }\n}\n\nfragment CommentInfo on Comment {\n  id\n  body\n  createdAt\n  parentId\n  author {\n    id\n    imageUrl(width: 200)\n    name\n    url\n    isBlocked\n    __typename\n  }\n  removedPerGuidelines\n  authorBadges\n  canReport\n  canDelete\n  canPin\n  hasFlaggings\n  deletedAuthor\n  deleted\n  sustained\n  pinnedAt\n  authorCanceledPledge\n  authorBacking {\n    backingUrl\n    id\n    __typename\n  }\n  __typename\n}\n\nfragment CommentReplies on Comment {\n  replies(last: 3, before: $replyCursor) {\n    totalCount\n    nodes {\n      ...CommentInfo\n      __typename\n    }\n    pageInfo {\n      startCursor\n      hasPreviousPage\n      __typename\n    }\n    __typename\n  }\n  __typename\n}',
        },
    ]

    headers = {
        'accept': '*/*',
        'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'content-type': 'application/json',
        'origin': 'https://www.kickstarter.com',
        'referer': url,
        'x-csrf-token': csrf_token,
    }
    session.headers.update(headers)
    response = session.post('https://www.kickstarter.com/graph', json=json_data)
    if response.status_code != 200:
        raise Exception(f'错误的响应码：{response.status_code}')

    response = response.json()

    if dedup.check_duplicate(url, json_data):
        print(f"重复记录，跳过写入，当前第:{page}页")
    else:
        with open(file_name, 'a', encoding='utf-8') as f:
            f.write(json.dumps(response) + '\n')

        comments = response[0]["data"]["commentable"]["comments"]["edges"]
        for comment in comments:
            main_comment = comment['node']['body']
            main_comment_author = (
                comment['node']['author'].get('name') if comment['node']['author'] else None
            )
            main_comment_create_time = comment['node']['createdAt']
            try:
                main_comment_create_time = datetime.fromtimestamp(
                    main_comment_create_time
                ).strftime('%Y-%m-%d %H:%M:%S')
            except:
                main_comment_create_time = main_comment_create_time
            reply_comments = []
            if comment['node'].get('replies', {}) and comment['node']['replies'].get('nodes', []):
                for reply in comment['node']['replies']['nodes']:
                    reply_comment_create_time = reply['createdAt']
                    try:
                        reply_comment_create_time = datetime.fromtimestamp(
                            reply_comment_create_time
                        ).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        reply_comment_create_time = reply_comment_create_time
                    reply_comments.append(
                        {
                            'reply_comment': reply["body"],
                            'reply_comment_author': reply['author'].get('name')
                            if reply['author']
                            else None,
                            'reply_comment_create_time': reply_comment_create_time,
                        }
                    )
            yield [main_comment, main_comment_author, main_comment_create_time, reply_comments]
        print(f'========= 第{page}页评论获取成功！ =========')
        dedup.add_record(url, json_data)
        if not PROXIES:
            time.sleep(random.randint(1, 3))

    if response[0]["data"]["commentable"]["comments"]["pageInfo"].get('hasNextPage'):
        print('检测到更多评论，正在执行翻页...')
        yield from get_comments(
            url,
            commentable_id,
            session,
            csrf_token,
            next_cursor=response[0]["data"]["commentable"]["comments"]["pageInfo"]['endCursor'],
            page=page + 1,
            file_name=file_name,
        )


def record_failed_task(url, project_id, msg):
    with open(FAILED_TASK_FILE, 'a', encoding='utf-8') as f:
        f.write(url.strip() + f' {project_id}' + ' >>> ' + msg + '\n')


def record_success_task(url):
    with open(SUCCESS_FLAG, 'a', encoding='utf-8') as f:
        f.write(url.strip() + '\n')


def process_failed_task(url, project_id):
    print('\n开始尝试抓取链接：', url)
    commentable_id = get_commentable_id(url)
    if not commentable_id:
        print(url, 'commentable_id提取失败，已跳过')
        return False
    try:
        target_csv_name = (
            str(project_id) + '_' + base64.b64decode(commentable_id).decode('utf-8') + '.csv'
        )
        target_txt_name = (
            str(project_id) + '_' + base64.b64decode(commentable_id).decode('utf-8') + '.txt'
        )
        target_csv_name = FOLDER_NAME + '/' + target_csv_name
        target_txt_name = FOLDER_NAME + '/' + target_txt_name
        for result in get_comments(url, commentable_id, file_name=target_txt_name):
            headers = ['链接', '评论内容', '评论作者', '评论时间', '子评论']
            result.insert(0, url)
            write_list_to_csv([result], headers, target_csv_name)
        return True
    except Exception as e:
        # raise
        print(url, '部分评论抓取失败 >>> ', e)
        return False


def failed_task_reexec(tasks=None, failed_task_file=''):
    """失败任务重试"""
    if tasks:
        pass
    elif failed_task_file:
        with open(failed_task_file, 'r', encoding='utf-8') as f:
            tasks = f.readlines()
    else:
        raise Exception('重试任务必须传递tasks或failed_task_file参数')

    # 要保留的行（未成功的）
    remaining_lines = []

    for task in tasks:
        url = task.strip(' ')[0]
        project_id = task.strip(' ')[1]
        if not process_failed_task(url, project_id):
            remaining_lines.append(task + '\n')  # 保留未成功的行

    # 重新写入文件，只写入未成功的行
    if remaining_lines:
        with open(failed_task_file, 'w', encoding='utf-8') as f:
            f.writelines(remaining_lines)
    else:
        Path(failed_task_file).unlink()


def main():
    global already_success_tasks
    global all_ready
    records = read_excel_to_dict_list(SOURCE_EXCEL)
    records.reverse()  # 倒序，从后往前抓，先抓评论多的

    # for data in tqdm(records, desc="kickstarter评论抓取"):
    for data in tqdm(records[4000:6000], desc="kickstarter评论抓取"):  # 第一批先抓4000
        url = data['Project link']
        if not all([url.strip() not in _ for _ in already_success_tasks]):
            print('当前链接已去重：', url)
            continue

        print('\n开始抓取链接：', url)
        session = random.choice(
            [
                requests.Session(
                    impersonate=random.choice(CURL_CFFI_BROWSER),
                    timeout=60,
                    proxies={'http': random.choice(PROXIES)},
                )
                if PROXIES
                else requests.Session(
                    impersonate=random.choice(CURL_CFFI_BROWSER),
                    timeout=60,
                ),
                # tls_client.Session(
                #     client_identifier=random.choice(TLS_CLIENT_BROWSER),
                #     random_tls_extension_order=True,
                # ),
            ]
        )
        commentable_result = get_commentable_id(url, session)
        if not commentable_result:
            print(url, 'commentable_id提取失败，已跳过')
            record_failed_task(url, str(data['Project ID']), 'commentable_id提取失败')
            time.sleep(60)
            continue
        try:
            commentable_id, csrf_token = commentable_result
            target_csv_name = (
                str(data['Project ID'])
                + '_'
                + base64.b64decode(commentable_id).decode('utf-8').split('-')[-1]
                + '.csv'
            )
            target_txt_name = (
                str(data['Project ID'])
                + '_'
                + base64.b64decode(commentable_id).decode('utf-8').split('-')[-1]
                + '.txt'
            )
            target_csv_name = FOLDER_NAME + '/' + target_csv_name
            target_txt_name = FOLDER_NAME + '/' + target_txt_name
            for result in get_comments(
                url, commentable_id, session, csrf_token, file_name=target_txt_name
            ):
                headers = ['链接', '评论内容', '评论作者', '评论时间', '子评论']
                result.insert(0, url)
                write_list_to_csv([result], headers, target_csv_name)

            record_success_task(url)
            already_success_tasks.append(url)
        except Exception as e:
            # raise
            print(url, '部分评论抓取失败 >>> ', e)
            record_failed_task(url, str(data['Project ID']), str(e))
            if '429' in str(e):
                print('检测到风控，5分钟后重试...')
                all_ready = False
                time.sleep(300)
            else:
                time.sleep(60)
        finally:
            if not PROXIES:
                time.sleep(random.randint(1, 3))


if __name__ == "__main__":
    with open(SUCCESS_FLAG, 'r', encoding='utf-8') as f1:
        already_success_tasks = f1.readlines()

    all_ready = True
    dedup = SQLiteDeduplicator()
    main()
    print('========== 全量抓取结束 ==========')
    for _ in range(50):
        if not all_ready:
            print(f'检测到遗漏项目，10秒后尝试补采，当前第【{_}/50】轮...')
            time.sleep(10)
            all_ready = True
            main()
